import re
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── 基礎票種對照（各隊通用）────────────────────────────
TICKET_MAP_BASE = {
    '全票':           '全票',
    '貴賓券':          '貴賓券',
    '公關票':          '公關票',
    '信用卡優惠票':     '信用卡優惠票',
    '啦啦隊票':        '啦啦隊票',
    '眷屬票':          '眷屬票',
    '內野優惠票':       '內野優惠票',
    '內野身心優惠票':   '身心優惠票',
    '外野身心優惠票':   '身心優惠票',
    '內野半票':         '半票',
    '外野半票':         '半票',
}

# ── 統一獅專屬票種 ─────────────────────────────────────
TICKET_MAP_LIONS = {
    **TICKET_MAP_BASE,
    '統一獅會員優惠票':    '會員優惠票(折100)',
    '統一卡優惠票':         '會員優惠票(折100)',
    '統一集團員工票':       '員工票',
    '統一超商優惠票':       '超商優惠票',
    '友好票':               '友好票',
    '內野記者票':           '記者票',
    '外野記者票':           '記者票',
}

# ── 樂天桃猿票種 ───────────────────────────────────────────
TICKET_MAP_RAKUTEN_WEEKDAY = {
    '尊猿門票':         '尊猿門票',
    '內野門票':         '內野門票',
    '內野全票':         '內野全票',
    '內野半票':         '內野半票',
    '外野門票':         '外野門票',
    '快樂炒門票':       '快樂炒門票',
    '大樂票':           '大樂票',
    '寵物樂園票':       '寵物樂園票',
    '身心障礙票.':      '身心障礙票',
    '身心障礙陪同票.':  '身心障礙陪同票',
    '貴賓票':           '貴賓票',
}

TICKET_MAP_RAKUTEN_HOLIDAY = {
    '尊猿門票．':       '尊猿門票',
    '內野門票．':       '內野門票',
    '內野全票．':       '內野全票',
    '內野半票．':       '內野半票',
    '外野門票．':       '外野門票',
    '快樂炒門票．':     '快樂炒門票',
    '大樂票．':         '大樂票',
    '寵物樂園票．':     '寵物樂園票',
    '身心障礙票.':      '身心障礙票',
    '身心障礙陪同票.':  '身心障礙陪同票',
}

# ── 各球隊 MAP 總表（新增球隊在此擴充）─────────────────
TICKET_MAPS = {
    '統一獅':           TICKET_MAP_LIONS,
    '樂天桃猿(平日)':   TICKET_MAP_RAKUTEN_WEEKDAY,
    '樂天桃猿(假日)':   TICKET_MAP_RAKUTEN_HOLIDAY,
    '通用':             TICKET_MAP_BASE,
}

# 向後相容
TICKET_MAP = TICKET_MAP_LIONS


def parse_ansource(file) -> dict:
    """讀取安源.xls，回傳 {date_label: DataFrame(area, ticket, price)}"""
    raw = pd.read_excel(file, sheet_name='票區票種資料(劃位)',
                        engine='xlrd', header=None)

    sections = []
    for i, row in raw.iterrows():
        for cell in row:
            if pd.notna(cell) and '演出日期' in str(cell):
                date_str = str(cell).replace('演出日期:', '').strip()
                sections.append((i, date_str))
                break

    header_rows = [
        i for i, row in raw.iterrows()
        if '票種編號' in row.values
    ]

    result = {}
    for idx, (sec_start, date_str) in enumerate(sections):
        data_header = next((hr for hr in header_rows if hr > sec_start), None)
        if data_header is None:
            continue
        data_start = data_header + 1
        data_end = sections[idx + 1][0] if idx + 1 < len(sections) else len(raw)

        df = raw.iloc[data_start:data_end].reset_index(drop=True)[[4, 8, 9]].copy()
        df.columns = ['area', 'ticket', 'price']
        df['area'] = df['area'].ffill()
        df = df.dropna(subset=['ticket'])
        df['price'] = pd.to_numeric(df['price'], errors='coerce')
        df = df.dropna(subset=['price'])
        df['area']   = df['area'].astype(str).str.strip()
        df['ticket'] = df['ticket'].astype(str).str.strip().str.rstrip('.')
        df['price']  = df['price'].astype(int)
        df = df[~df['ticket'].isin(['票種', 'nan'])]
        df = df.drop_duplicates()
        result[date_str] = df

    return result


def parse_vendor(file) -> dict:
    """讀取廠商.xlsx，回傳含「票價」的頁籤 {sheet_name: DataFrame(area, ticket_vs, ticket_std, price)}"""
    xl = pd.ExcelFile(file, engine='openpyxl')
    result = {}
    for sheet in xl.sheet_names:
        if '票價' not in sheet:
            continue
        raw = pd.read_excel(file, sheet_name=sheet,
                            engine='openpyxl', header=None)
        # 動態找出 row 2 中 col 2 起有值的票種欄位
        ticket_row = raw.iloc[2]
        ticket_cols = [c for c in range(2, len(ticket_row))
                       if pd.notna(ticket_row[c]) and str(ticket_row[c]).strip() != '']
        ticket_types = [ticket_row[c] for c in ticket_cols]

        df_wide = raw.iloc[3:].copy()
        available_cols = [c for c in [1] + ticket_cols if c in df_wide.columns]
        df_wide = df_wide[available_cols].copy()
        df_wide.columns = ['area'] + ticket_types[:len(available_cols) - 1]
        df_wide = df_wide.dropna(subset=['area'])
        df_wide['area'] = df_wide['area'].astype(str).str.strip()

        df = df_wide.melt(id_vars=['area'], var_name='ticket_vs', value_name='price')
        df = df[df['price'] != '-'].dropna(subset=['price'])
        df['price'] = pd.to_numeric(df['price'], errors='coerce')
        df = df.dropna(subset=['price'])
        df['price'] = df['price'].astype(int)
        # 過濾 price=0 的隱藏票種（標記為「不顯示」者售價為 0，不應納入比對）
        df = df[df['price'] > 0]
        df['ticket_std'] = (df['ticket_vs']
                            .str.replace(r'\(不顯示\)', '', regex=True)
                            .str.strip())
        df = df.drop_duplicates()
        result[sheet] = df
    return result


def _extract_price(s) -> int | None:
    """從字串中提取第一個整數票價，無效值回傳 None。"""
    s = str(s).strip() if s is not None and not (isinstance(s, float) and pd.isna(s)) else ''
    if s in ('X', 'x', '-', 'nan', 'NaN', ''):
        return None
    m = re.search(r'\d+', s)
    return int(m.group()) if m else None


def _pairs(price_cell, ticket_cell, last_ticket: str | None = None) -> list[tuple[str, int]]:
    """解析多行複合儲存格，回傳 [(ticket_name, price), ...]。"""
    p = str(price_cell).strip() if pd.notna(price_cell) else ''
    t_raw = str(ticket_cell).strip() if pd.notna(ticket_cell) and str(ticket_cell).strip() not in ('nan', 'NaN', '') else ''
    pl = [l.strip() for l in p.split('\n') if l.strip()]
    tl = [l.strip() for l in t_raw.split('\n') if l.strip()]
    if not tl and last_ticket:
        tl = [last_ticket]
    out = []
    for i, pline in enumerate(pl):
        price = _extract_price(pline)
        if price is None:
            continue
        if tl:
            ticket = tl[i] if i < len(tl) else tl[0]
        else:
            ticket = ('內野全票' if '全票' in pline else ('內野半票' if '半票' in pline else None))
        if ticket:
            out.append((ticket, price))
    return out


def _canon_area_rakuten_vendor(main_area, sub_area) -> str:
    """廠商 Excel 區域欄位 → 統一區域名稱。"""
    m = re.sub(r'\n.*', '', str(main_area or '')).strip()
    m = re.sub(r'[A-Za-z].*', '', m).strip()
    m = re.sub(r'[※*].*', '', m).strip()
    s = str(sub_area or '').strip()
    if '尊猿席' in m:
        return '尊猿席'
    if m in ('東西下', '東下', '西下'):
        if re.search(r'^[AB]', s):
            return '東西下(AB)'
        if 'DE' in s and '前' in s:
            return '東西下(DE熱)'
        if 'CFG' in s or ('13' in s and '後' in s):
            return '東西下(CFG)'
        if re.search(r'^[H-M]', s):
            return '東下' if m != '西下' else '西下'
        if s == 'R':
            return '東西下(R)'
        return m
    if m == '東西上':
        return '東西上'
    if '外野' in m and '大樂' not in m:
        return '外野'
    if '大樂' in m:
        return '大樂放鬆區'
    if '快樂炒' in m:
        return '快樂炒'
    if '團猿席' in m:
        return '東下團猿席'
    if '上層' in m:
        return '上層團猿席'
    if '寵物' in m:
        return '寵物樂園'
    return m or '?'


def _canon_area_rakuten_ay(area: str) -> str | None:
    """安源 Excel 區域欄位 → 統一區域名稱（回傳 None 表示跳過此列）。"""
    a = str(area or '').strip()
    if '尊猿席' in a:
        return '尊猿席'
    if '熱力應援' in a:
        return '東西下(DE熱)'
    if re.fullmatch(r'[東西]下[AB]區', a):
        return '東西下(AB)'
    if re.fullmatch(r'[東西]下[CDEFG]區', a):
        return '東西下(CFG)'
    if re.fullmatch(r'東下[H-M]區', a):
        return '東下'
    if re.fullmatch(r'西下[H-K]區', a):
        return '西下'
    if re.fullmatch(r'[東西]上[A-M]區', a):
        return '東西上'
    if re.search(r'[東西]R\d', a) or 'DAZN' in a:
        return '東西下(R)'
    if '右外野' in a or '左外野' in a:
        return '外野'
    if '外野' in a:
        return '外野'
    if '大樂' in a:
        return '大樂放鬆區'
    if '快樂炒' in a:
        return '快樂炒'
    if '輪椅' in a:
        return None  # 輪椅席不納入比對
    return a


def parse_vendor_rakuten(file) -> dict:
    """讀取樂天廠商 .xlsx，回傳 {sheet_name: DataFrame(area, ticket_vs, ticket_std, price)}"""
    xl = pd.ExcelFile(file, engine='openpyxl')
    result = {}
    for sheet in xl.sheet_names:
        if '票價' not in sheet:
            continue
        raw = pd.read_excel(file, sheet_name=sheet, engine='openpyxl', header=None)

        rows = []
        last_ticket = None
        for _, row in raw.iterrows():
            main_area = row.iloc[0] if len(row) > 0 else None
            sub_area  = row.iloc[1] if len(row) > 1 else None
            price_cell  = row.iloc[2] if len(row) > 2 else None
            ticket_cell = row.iloc[3] if len(row) > 3 else None

            if pd.isna(main_area) and pd.isna(sub_area):
                continue
            if _extract_price(price_cell) is None and not str(price_cell).strip():
                continue

            canon_area = _canon_area_rakuten_vendor(main_area, sub_area)
            pairs = _pairs(price_cell, ticket_cell, last_ticket)
            if pairs:
                last_ticket = pairs[-1][0]
            for ticket, price in pairs:
                rows.append({'area': canon_area, 'ticket_vs': ticket,
                             'ticket_std': ticket, 'price': price})

        if rows:
            df = pd.DataFrame(rows)
            df['price'] = df['price'].astype(int)
            df = df[df['price'] > 0].drop_duplicates()
            result[sheet] = df
    return result


def compare(
    df_ay: pd.DataFrame,
    df_vs: pd.DataFrame,
    ticket_map: dict | None = None,
    ay_area_map: dict | None = None,
) -> dict:
    """比對安源與廠商資料，回傳分類結果 dict。

    Args:
        df_ay:        安源 DataFrame，欄位: area, ticket, price
        df_vs:        廠商 DataFrame，欄位: area, ticket_vs, ticket_std, price
        ticket_map:   {安源票種 → ticket_std}；None 時使用預設 TICKET_MAP
        ay_area_map:  {安源area → 廠商area key}；None 時不做區域轉換
    """
    if ticket_map is None:
        ticket_map = TICKET_MAP

    df_ay = df_ay.copy()
    df_vs = df_vs.copy()

    # 安源票種標準化
    df_ay['ticket_std'] = df_ay['ticket'].map(ticket_map)

    # 安源區域對應（支援 dict 或 callable；callable 回傳 None 表示跳過）
    if ay_area_map:
        if callable(ay_area_map):
            df_ay['area'] = df_ay['area'].map(ay_area_map)
            df_ay = df_ay[df_ay['area'].notna()]
        else:
            df_ay['area'] = df_ay['area'].map(lambda a: ay_area_map.get(a, a))

    ay_m = (df_ay[df_ay['ticket_std'].notna()]
            [['area', 'ticket', 'ticket_std', 'price']]
            .rename(columns={'price': 'price_ay', 'ticket': 'ticket_ay'}))

    vs_m = (df_vs[['area', 'ticket_vs', 'ticket_std', 'price']]
            .rename(columns={'price': 'price_vs'}))

    # Ensure ticket_std dtype is consistent to avoid merge errors on empty DataFrames
    ay_m = ay_m.copy(); ay_m['ticket_std'] = ay_m['ticket_std'].astype(object)
    vs_m = vs_m.copy(); vs_m['ticket_std'] = vs_m['ticket_std'].astype(object)

    merged = pd.merge(ay_m, vs_m, on=['area', 'ticket_std'],
                      how='outer', indicator=True)

    both     = merged[merged['_merge'] == 'both'].drop(columns='_merge').copy()
    only_ay  = merged[merged['_merge'] == 'left_only'][['area', 'ticket_ay', 'ticket_std', 'price_ay']].copy()
    only_vs  = merged[merged['_merge'] == 'right_only'][['area', 'ticket_vs', 'ticket_std', 'price_vs']].copy()

    return {
        'price_ok':   both[both['price_ay'] == both['price_vs']].copy(),
        'price_diff': both[both['price_ay'] != both['price_vs']].copy(),
        'only_ay':    only_ay,
        'only_vs':    only_vs,
        'unmapped':   df_ay[df_ay['ticket_std'].isna()][['area', 'ticket', 'price']].copy(),
    }


def generate_excel(result: dict) -> bytes:
    """將比對結果產出為 Excel bytes（in-memory，不落地）"""
    wb = Workbook()
    GREEN  = PatternFill('solid', fgColor='C6EFCE')
    RED    = PatternFill('solid', fgColor='FFC7CE')
    ORANGE = PatternFill('solid', fgColor='FFEB9C')
    BLUE   = PatternFill('solid', fgColor='BDD7EE')
    GRAY   = PatternFill('solid', fgColor='D9D9D9')
    HFILL  = PatternFill('solid', fgColor='4472C4')

    def write_sheet(ws, df, cols, col_names, fill, title, col_widths):
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 1
        ws.cell(r, 1, title).font = Font(bold=True, size=11, color='FFFFFF')
        ws.cell(r, 1).fill = HFILL
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=len(cols))
        ws.cell(r, 1).alignment = Alignment(horizontal='center')
        r += 1
        for c, name in enumerate(col_names, 1):
            cell = ws.cell(r, c, name)
            cell.font = Font(bold=True)
            cell.fill = GRAY
            cell.alignment = Alignment(horizontal='center')
        r += 1
        for row in df[cols].itertuples(index=False):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val).fill = fill
            r += 1

    price_ok   = result['price_ok']
    price_diff = result['price_diff']
    only_ay    = result['only_ay']
    only_vs    = result['only_vs']
    unmapped   = result['unmapped']

    # 摘要
    ws0 = wb.active
    ws0.title = '摘要'
    ws0.column_dimensions['A'].width = 38
    ws0.column_dimensions['B'].width = 12
    for r, (label, count, fill) in enumerate([
        ('✅ 票價完全相符',      len(price_ok),   GREEN),
        ('❌ 票價不符',          len(price_diff), RED),
        ('⚠️ 廠商有、安源沒有',  len(only_vs),    ORANGE),
        ('⚠️ 安源有、廠商沒有',  len(only_ay),    ORANGE),
        ('ℹ️ 安源票種無對應廠商', len(unmapped),  BLUE),
    ], 1):
        ws0.cell(r, 1, label).fill = fill
        ws0.cell(r, 2, count).fill = fill

    if len(price_diff) > 0:
        ws = wb.create_sheet('❌票價不符')
        write_sheet(ws, price_diff.sort_values(['ticket_std', 'area']),
            ['area', 'ticket_vs', 'ticket_ay', 'price_vs', 'price_ay'],
            ['區域名稱', '廠商票種', '安源票種', '廠商票價', '安源票價'],
            RED, f'❌ 票價不符 ({len(price_diff)}筆)', [42, 20, 20, 12, 12])

    if len(only_vs) > 0:
        ws = wb.create_sheet('⚠️廠商有安源無')
        write_sheet(ws, only_vs.sort_values(['ticket_std', 'area']),
            ['area', 'ticket_vs', 'price_vs'],
            ['區域名稱', '廠商票種', '廠商票價'],
            ORANGE, f'⚠️ 廠商有、安源沒有 ({len(only_vs)}筆)', [42, 22, 12])

    if len(only_ay) > 0:
        ws = wb.create_sheet('⚠️安源有廠商無')
        write_sheet(ws, only_ay.sort_values(['area', 'ticket_ay']),
            ['area', 'ticket_ay', 'price_ay'],
            ['區域名稱', '安源票種', '安源票價'],
            ORANGE, f'⚠️ 安源有、廠商沒有 ({len(only_ay)}筆)', [42, 22, 12])

    if len(unmapped) > 0:
        ws = wb.create_sheet('ℹ️安源無對應')
        write_sheet(ws, unmapped.sort_values(['ticket', 'area']),
            ['area', 'ticket', 'price'],
            ['區域名稱', '安源票種', '安源票價'],
            BLUE, f'ℹ️ 安源票種無對應廠商 ({len(unmapped)}筆)', [42, 22, 12])

    if len(price_ok) > 0:
        ws = wb.create_sheet('✅完全相符')
        ok_display = (price_ok.sort_values(['ticket_std', 'area'])
                              [['area', 'ticket_vs', 'ticket_ay', 'price_vs']]
                              .copy()
                              .rename(columns={'price_vs': 'price'}))
        write_sheet(ws, ok_display,
            ['area', 'ticket_vs', 'ticket_ay', 'price'],
            ['區域名稱', '廠商票種', '安源票種', '票價'],
            GREEN, f'✅ 票價完全相符 ({len(price_ok)}筆)', [42, 22, 22, 12])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
